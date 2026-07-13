"""
Agent DCF & Growth Analysis — Streamlit web app
=================================================
Déploiement : Streamlit Community Cloud (gratuit) via GitHub.
Usage local : streamlit run app.py
"""

import io
from datetime import datetime

import numpy as np
import pandas as pd
import streamlit as st
import yfinance as yf
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Agent DCF", page_icon="📊", layout="wide")


# ---------------------------------------------------------------------------
# 1. INGESTION
# ---------------------------------------------------------------------------

@st.cache_data(ttl=3600)
def fetch_company_data(ticker: str) -> dict:
    t = yf.Ticker(ticker)
    info = t.info or {}
    news = t.news or []
    return {
        "ticker": ticker,
        "name": info.get("longName", ticker),
        "sector": info.get("sector"),
        "price": info.get("currentPrice") or info.get("regularMarketPrice"),
        "shares_out": info.get("sharesOutstanding"),
        "market_cap": info.get("marketCap"),
        "beta": info.get("beta"),
        "total_debt": info.get("totalDebt"),
        "total_cash": info.get("totalCash"),
        "ebitda": info.get("ebitda"),
        "revenue": info.get("totalRevenue"),
        "analyst_target": info.get("targetMeanPrice"),
        "news": [n.get("title") for n in news[:8] if n.get("title")],
    }


# ---------------------------------------------------------------------------
# 2. ANALYSE — DCF déterministe
# ---------------------------------------------------------------------------

def compute_wacc(market_cap, total_debt, beta, rf, erp, pretax_cost_debt, tax_rate):
    market_cap = market_cap or 0
    total_debt = total_debt or 0
    beta = beta if beta is not None else 1.0
    cost_equity = rf + beta * erp
    cost_debt_at = pretax_cost_debt * (1 - tax_rate)
    V = (market_cap + total_debt) or 1
    we, wd = market_cap / V, total_debt / V
    wacc = we * cost_equity + wd * cost_debt_at
    return {"wacc": wacc, "cost_equity": cost_equity, "cost_debt_at": cost_debt_at,
            "we": we, "wd": wd}


def project_fcf(ebitda_base, da, tax_rate, annual_growth, capex_pct_rev, revenue_base, nwc_change, years):
    ebitda = ebitda_base
    revenue = revenue_base
    rows = []
    for y in years:
        ebitda *= (1 + annual_growth)
        revenue *= (1 + annual_growth)
        ebit = ebitda - da
        nopat = ebit * (1 - tax_rate)
        capex = revenue * capex_pct_rev
        fcf = nopat + da - capex - nwc_change
        rows.append({"year": y, "ebitda": ebitda, "capex": capex, "fcf": fcf})
    return pd.DataFrame(rows)


def dcf_valuation(fcf_df, wacc, g_terminal, net_debt, shares_out):
    n = len(fcf_df)
    pv_fcf = sum(row.fcf / (1 + wacc) ** (i + 1) for i, row in enumerate(fcf_df.itertuples()))
    fcf_terminal = fcf_df.iloc[-1].fcf * (1 + g_terminal)
    tv = fcf_terminal / (wacc - g_terminal)
    pv_tv = tv / (1 + wacc) ** n
    ev = pv_fcf + pv_tv
    equity_value = ev - net_debt
    value_per_share = equity_value / shares_out if shares_out else None
    return {"pv_fcf": pv_fcf, "pv_tv": pv_tv, "ev": ev,
            "equity_value": equity_value, "value_per_share": value_per_share}


def sensitivity_grid(fcf_df, wacc, net_debt, shares_out, wacc_deltas, g_values):
    grid = np.zeros((len(wacc_deltas), len(g_values)))
    for i, d in enumerate(wacc_deltas):
        for j, g in enumerate(g_values):
            v = dcf_valuation(fcf_df, wacc + d, g, net_debt, shares_out)["value_per_share"]
            grid[i, j] = v if v is not None else np.nan
    return grid


# ---------------------------------------------------------------------------
# 3. EXPORT EXCEL (en mémoire, pour le bouton de téléchargement)
# ---------------------------------------------------------------------------

def build_excel(company, wacc_result, fcf_df, dcf_result, sens_grid, wacc_deltas, g_values):
    wb = Workbook()
    ws = wb.active
    ws.title = "DCF Model"
    blue = Font(color="0000FF")
    bold = Font(bold=True)
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    ws["A1"] = f"{company['name']} ({company['ticker']}) — DCF model"
    ws["A1"].font = Font(bold=True, size=13)

    ws["A3"] = "Current price"; ws["B3"] = company["price"]; ws["B3"].font = blue
    ws["A4"] = "Shares outstanding"; ws["B4"] = company["shares_out"]; ws["B4"].font = blue
    ws["A5"] = "WACC"; ws["B5"] = wacc_result["wacc"]; ws["B5"].number_format = "0.0%"
    ws["A6"] = "Implied value / share"
    ws["B6"] = dcf_result["value_per_share"]
    ws["B6"].font = bold
    ws["B6"].fill = yellow
    ws["B6"].number_format = '$#,##0.00'

    ws["A8"] = "Year"; ws["B8"] = "EBITDA"; ws["C8"] = "Capex"; ws["D8"] = "FCF"
    for i, row in fcf_df.iterrows():
        r = 9 + i
        ws.cell(r, 1, int(row.year))
        ws.cell(r, 2, round(row.ebitda, 1))
        ws.cell(r, 3, round(row.capex, 1))
        ws.cell(r, 4, round(row.fcf, 1))

    ws2 = wb.create_sheet("Sensitivity")
    ws2.cell(1, 1, "WACC \\ g")
    for j, g in enumerate(g_values):
        ws2.cell(1, 2 + j, f"{g:.1%}")
    for i, d in enumerate(wacc_deltas):
        ws2.cell(2 + i, 1, f"{wacc_result['wacc'] + d:.2%}")
        for j, g in enumerate(g_values):
            val = sens_grid[i, j]
            ws2.cell(2 + i, 2 + j, round(val, 2) if not np.isnan(val) else None)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# 4. INTERFACE
# ---------------------------------------------------------------------------

st.title("📊 Agent DCF — analyse de valorisation")
st.caption("Entre un ticker, ajuste les hypothèses, obtiens un DCF complet avec dashboard et export Excel.")

with st.sidebar:
    st.header("Paramètres")
    ticker = st.text_input("Ticker", value="CAT").upper().strip()
    st.subheader("Hypothèses WACC")
    rf = st.slider("Taux sans risque", 0.02, 0.06, 0.043, 0.001, format="%.3f")
    erp = st.slider("Prime de risque marché", 0.03, 0.07, 0.05, 0.001, format="%.3f")
    beta_override = st.number_input("Beta (0 = utiliser celui de yfinance)", 0.0, 3.0, 0.0, 0.05)
    pretax_cost_debt = st.slider("Coût de la dette (avant impôt)", 0.02, 0.10, 0.055, 0.001, format="%.3f")
    tax_rate = st.slider("Taux d'imposition", 0.10, 0.35, 0.25, 0.01)
    st.subheader("Hypothèses de croissance")
    annual_growth = st.slider("Croissance annuelle EBITDA/CA", -0.05, 0.15, 0.03, 0.01)
    capex_pct_rev = st.slider("Capex (% du chiffre d'affaires)", 0.02, 0.15, 0.06, 0.005)
    n_years = st.slider("Horizon explicite (années)", 3, 10, 5)
    g_terminal = st.slider("Croissance terminale", 0.0, 0.04, 0.02, 0.005)
    run_btn = st.button("Lancer l'analyse", type="primary", use_container_width=True)

if run_btn and ticker:
    with st.spinner(f"Récupération des données pour {ticker}..."):
        try:
            company = fetch_company_data(ticker)
        except Exception as e:
            st.error(f"Erreur lors de la récupération des données : {e}")
            st.stop()

    if not company.get("price") or not company.get("shares_out"):
        st.warning("Données incomplètes pour ce ticker (prix ou actions en circulation manquants). "
                   "Certains petits caps ou tickers non-US sont mal couverts par yfinance.")
        st.stop()

    beta = beta_override if beta_override > 0 else (company["beta"] or 1.2)
    wacc_result = compute_wacc(company["market_cap"], company["total_debt"], beta,
                                rf, erp, pretax_cost_debt, tax_rate)

    years = list(range(datetime.now().year + 1, datetime.now().year + 1 + n_years))
    da = (company["ebitda"] or 0) * 0.35
    fcf_df = project_fcf(
        ebitda_base=company["ebitda"] or 1,
        da=da,
        tax_rate=tax_rate,
        annual_growth=annual_growth,
        capex_pct_rev=capex_pct_rev,
        revenue_base=company["revenue"] or (company["ebitda"] or 1) * 5,
        nwc_change=0,
        years=years,
    )
    net_debt = (company["total_debt"] or 0) - (company["total_cash"] or 0)
    dcf_result = dcf_valuation(fcf_df, wacc_result["wacc"], g_terminal, net_debt, company["shares_out"])

    wacc_deltas = [-0.01, -0.005, 0, 0.005, 0.01]
    g_values = [max(0.0, g_terminal - 0.01), g_terminal - 0.005, g_terminal, g_terminal + 0.005, g_terminal + 0.01]
    sens_grid = sensitivity_grid(fcf_df, wacc_result["wacc"], net_debt, company["shares_out"], wacc_deltas, g_values)

    st.subheader(f"{company['name']} ({company['ticker']})")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Cours actuel", f"${company['price']:.2f}")
    c2.metric("Valeur DCF / action", f"${dcf_result['value_per_share']:.2f}" if dcf_result['value_per_share'] else "N/A")
    upside = (dcf_result['value_per_share'] / company['price'] - 1) if dcf_result['value_per_share'] else None
    c3.metric("Écart DCF vs marché", f"{upside:.1%}" if upside is not None else "N/A")
    c4.metric("WACC", f"{wacc_result['wacc']:.2%}")

    col1, col2 = st.columns([1.3, 1])
    with col1:
        fig = go.Figure()
        fig.add_bar(x=fcf_df["year"], y=fcf_df["ebitda"], name="EBITDA")
        fig.add_scatter(x=fcf_df["year"], y=fcf_df["fcf"], name="FCF", mode="lines+markers")
        fig.update_layout(title="Projection EBITDA & FCF", height=350, margin=dict(t=40, b=20))
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        fig2 = go.Figure(data=[go.Pie(
            labels=["Fonds propres", "Dette"],
            values=[wacc_result["we"], wacc_result["wd"]],
            hole=0.6,
        )])
        fig2.update_layout(title=f"WACC = {wacc_result['wacc']:.2%}", height=350, margin=dict(t=40, b=20))
        st.plotly_chart(fig2, use_container_width=True)"""
Agent DCF & Growth Analysis — Streamlit web app
=================================================
Déploiement : Streamlit Community Cloud (gratuit) via GitHub.
Usage local : streamlit run app.py
"""

import io
from datetime import datetime

import numpy as np
import pandas as pd
import streamlit as st
import yfinance as yf
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from curl_cffi import requests as curl_requests
    _SESSION = curl_requests.Session(impersonate="chrome")
except ImportError:
    _SESSION = None

st.set_page_config(page_title="Agent DCF", page_icon="📊", layout="wide")


# ---------------------------------------------------------------------------
# 1. INGESTION
# ---------------------------------------------------------------------------

@st.cache_data(ttl=3600)
def fetch_company_data(ticker: str) -> dict:
    t = yf.Ticker(ticker, session=_SESSION) if _SESSION else yf.Ticker(ticker)
    info = t.info or {}
    news = t.news or []
    return {
        "ticker": ticker,
        "name": info.get("longName", ticker),
        "sector": info.get("sector"),
        "price": info.get("currentPrice") or info.get("regularMarketPrice"),
        "shares_out": info.get("sharesOutstanding"),
        "market_cap": info.get("marketCap"),
        "beta": info.get("beta"),
        "total_debt": info.get("totalDebt"),
        "total_cash": info.get("totalCash"),
        "ebitda": info.get("ebitda"),
        "revenue": info.get("totalRevenue"),
        "analyst_target": info.get("targetMeanPrice"),
        "news": [n.get("title") for n in news[:8] if n.get("title")],
    }


# ---------------------------------------------------------------------------
# 2. ANALYSE — DCF déterministe
# ---------------------------------------------------------------------------

def compute_wacc(market_cap, total_debt, beta, rf, erp, pretax_cost_debt, tax_rate):
    market_cap = market_cap or 0
    total_debt = total_debt or 0
    beta = beta if beta is not None else 1.0
    cost_equity = rf + beta * erp
    cost_debt_at = pretax_cost_debt * (1 - tax_rate)
    V = (market_cap + total_debt) or 1
    we, wd = market_cap / V, total_debt / V
    wacc = we * cost_equity + wd * cost_debt_at
    return {"wacc": wacc, "cost_equity": cost_equity, "cost_debt_at": cost_debt_at,
            "we": we, "wd": wd}


def project_fcf(ebitda_base, da, tax_rate, annual_growth, capex_pct_rev, revenue_base, nwc_change, years):
    ebitda = ebitda_base
    revenue = revenue_base
    rows = []
    for y in years:
        ebitda *= (1 + annual_growth)
        revenue *= (1 + annual_growth)
        ebit = ebitda - da
        nopat = ebit * (1 - tax_rate)
        capex = revenue * capex_pct_rev
        fcf = nopat + da - capex - nwc_change
        rows.append({"year": y, "ebitda": ebitda, "capex": capex, "fcf": fcf})
    return pd.DataFrame(rows)


def dcf_valuation(fcf_df, wacc, g_terminal, net_debt, shares_out):
    n = len(fcf_df)
    pv_fcf = sum(row.fcf / (1 + wacc) ** (i + 1) for i, row in enumerate(fcf_df.itertuples()))
    fcf_terminal = fcf_df.iloc[-1].fcf * (1 + g_terminal)
    tv = fcf_terminal / (wacc - g_terminal)
    pv_tv = tv / (1 + wacc) ** n
    ev = pv_fcf + pv_tv
    equity_value = ev - net_debt
    value_per_share = equity_value / shares_out if shares_out else None
    return {"pv_fcf": pv_fcf, "pv_tv": pv_tv, "ev": ev,
            "equity_value": equity_value, "value_per_share": value_per_share}


def sensitivity_grid(fcf_df, wacc, net_debt, shares_out, wacc_deltas, g_values):
    grid = np.zeros((len(wacc_deltas), len(g_values)))
    for i, d in enumerate(wacc_deltas):
        for j, g in enumerate(g_values):
            v = dcf_valuation(fcf_df, wacc + d, g, net_debt, shares_out)["value_per_share"]
            grid[i, j] = v if v is not None else np.nan
    return grid


# ---------------------------------------------------------------------------
# 3. EXPORT EXCEL (en mémoire, pour le bouton de téléchargement)
# ---------------------------------------------------------------------------

def build_excel(company, wacc_result, fcf_df, dcf_result, sens_grid, wacc_deltas, g_values):
    wb = Workbook()
    ws = wb.active
    ws.title = "DCF Model"
    blue = Font(color="0000FF")
    bold = Font(bold=True)
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    ws["A1"] = f"{company['name']} ({company['ticker']}) — DCF model"
    ws["A1"].font = Font(bold=True, size=13)

    ws["A3"] = "Current price"; ws["B3"] = company["price"]; ws["B3"].font = blue
    ws["A4"] = "Shares outstanding"; ws["B4"] = company["shares_out"]; ws["B4"].font = blue
    ws["A5"] = "WACC"; ws["B5"] = wacc_result["wacc"]; ws["B5"].number_format = "0.0%"
    ws["A6"] = "Implied value / share"
    ws["B6"] = dcf_result["value_per_share"]
    ws["B6"].font = bold
    ws["B6"].fill = yellow
    ws["B6"].number_format = '$#,##0.00'

    ws["A8"] = "Year"; ws["B8"] = "EBITDA"; ws["C8"] = "Capex"; ws["D8"] = "FCF"
    for i, row in fcf_df.iterrows():
        r = 9 + i
        ws.cell(r, 1, int(row.year))
        ws.cell(r, 2, round(row.ebitda, 1))
        ws.cell(r, 3, round(row.capex, 1))
        ws.cell(r, 4, round(row.fcf, 1))

    ws2 = wb.create_sheet("Sensitivity")
    ws2.cell(1, 1, "WACC \\ g")
    for j, g in enumerate(g_values):
        ws2.cell(1, 2 + j, f"{g:.1%}")
    for i, d in enumerate(wacc_deltas):
        ws2.cell(2 + i, 1, f"{wacc_result['wacc'] + d:.2%}")
        for j, g in enumerate(g_values):
            val = sens_grid[i, j]
            ws2.cell(2 + i, 2 + j, round(val, 2) if not np.isnan(val) else None)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# 4. INTERFACE
# ---------------------------------------------------------------------------

st.title("📊 Agent DCF — analyse de valorisation")
st.caption("Entre un ticker, ajuste les hypothèses, obtiens un DCF complet avec dashboard et export Excel.")

with st.sidebar:
    st.header("Paramètres")
    ticker = st.text_input("Ticker", value="CAT").upper().strip()
    st.subheader("Hypothèses WACC")
    rf = st.slider("Taux sans risque", 0.02, 0.06, 0.043, 0.001, format="%.3f")
    erp = st.slider("Prime de risque marché", 0.03, 0.07, 0.05, 0.001, format="%.3f")
    beta_override = st.number_input("Beta (0 = utiliser celui de yfinance)", 0.0, 3.0, 0.0, 0.05)
    pretax_cost_debt = st.slider("Coût de la dette (avant impôt)", 0.02, 0.10, 0.055, 0.001, format="%.3f")
    tax_rate = st.slider("Taux d'imposition", 0.10, 0.35, 0.25, 0.01)
    st.subheader("Hypothèses de croissance")
    annual_growth = st.slider("Croissance annuelle EBITDA/CA", -0.05, 0.15, 0.03, 0.01)
    capex_pct_rev = st.slider("Capex (% du chiffre d'affaires)", 0.02, 0.15, 0.06, 0.005)
    n_years = st.slider("Horizon explicite (années)", 3, 10, 5)
    g_terminal = st.slider("Croissance terminale", 0.0, 0.04, 0.02, 0.005)
    run_btn = st.button("Lancer l'analyse", type="primary", use_container_width=True)

if run_btn and ticker:
    with st.spinner(f"Récupération des données pour {ticker}..."):
        try:
            company = fetch_company_data(ticker)
        except Exception as e:
            st.error(f"Erreur lors de la récupération des données : {e}")
            st.stop()

    if not company.get("price") or not company.get("shares_out"):
        st.warning("Données incomplètes pour ce ticker (prix ou actions en circulation manquants). "
                   "Certains petits caps ou tickers non-US sont mal couverts par yfinance.")
        st.stop()

    beta = beta_override if beta_override > 0 else (company["beta"] or 1.2)
    wacc_result = compute_wacc(company["market_cap"], company["total_debt"], beta,
                                rf, erp, pretax_cost_debt, tax_rate)

    years = list(range(datetime.now().year + 1, datetime.now().year + 1 + n_years))
    da = (company["ebitda"] or 0) * 0.35
    fcf_df = project_fcf(
        ebitda_base=company["ebitda"] or 1,
        da=da,
        tax_rate=tax_rate,
        annual_growth=annual_growth,
        capex_pct_rev=capex_pct_rev,
        revenue_base=company["revenue"] or (company["ebitda"] or 1) * 5,
        nwc_change=0,
        years=years,
    )
    net_debt = (company["total_debt"] or 0) - (company["total_cash"] or 0)
    dcf_result = dcf_valuation(fcf_df, wacc_result["wacc"], g_terminal, net_debt, company["shares_out"])

    wacc_deltas = [-0.01, -0.005, 0, 0.005, 0.01]
    g_values = [max(0.0, g_terminal - 0.01), g_terminal - 0.005, g_terminal, g_terminal + 0.005, g_terminal + 0.01]
    sens_grid = sensitivity_grid(fcf_df, wacc_result["wacc"], net_debt, company["shares_out"], wacc_deltas, g_values)

    st.subheader(f"{company['name']} ({company['ticker']})")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Cours actuel", f"${company['price']:.2f}")
    c2.metric("Valeur DCF / action", f"${dcf_result['value_per_share']:.2f}" if dcf_result['value_per_share'] else "N/A")
    upside = (dcf_result['value_per_share'] / company['price'] - 1) if dcf_result['value_per_share'] else None
    c3.metric("Écart DCF vs marché", f"{upside:.1%}" if upside is not None else "N/A")
    c4.metric("WACC", f"{wacc_result['wacc']:.2%}")

    col1, col2 = st.columns([1.3, 1])
    with col1:
        fig = go.Figure()
        fig.add_bar(x=fcf_df["year"], y=fcf_df["ebitda"], name="EBITDA")
        fig.add_scatter(x=fcf_df["year"], y=fcf_df["fcf"], name="FCF", mode="lines+markers")
        fig.update_layout(title="Projection EBITDA & FCF", height=350, margin=dict(t=40, b=20))
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        fig2 = go.Figure(data=[go.Pie(
            labels=["Fonds propres", "Dette"],
            values=[wacc_result["we"], wacc_result["wd"]],
            hole=0.6,
        )])
        fig2.update_layout(title=f"WACC = {wacc_result['wacc']:.2%}", height=350, margin=dict(t=40, b=20))
        st.plotly_chart(fig2, use_container_width=True)

    st.subheader("Sensibilité — valeur par action ($)")
    sens_df = pd.DataFrame(
        sens_grid,
        index=[f"{wacc_result['wacc'] + d:.2%}" for d in wacc_deltas],
        columns=[f"{g:.1%}" for g in g_values],
    )
    st.dataframe(sens_df.style.format("{:.2f}").background_gradient(cmap="RdYlBu_r", axis=None),
                 use_container_width=True)

    if company["news"]:
        st.subheader("Actualités récentes")
        for title in company["news"]:
            st.write(f"- {title}")

    excel_buf = build_excel(company, wacc_result, fcf_df, dcf_result, sens_grid, wacc_deltas, g_values)
    st.download_button(
        "📥 Télécharger le modèle Excel",
        data=excel_buf,
        file_name=f"{ticker}_DCF_model.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

else:
    st.info("Renseigne un ticker dans la barre latérale et clique sur \"Lancer l'analyse\".")


    st.subheader("Sensibilité — valeur par action ($)")
    sens_df = pd.DataFrame(
        sens_grid,
        index=[f"{wacc_result['wacc'] + d:.2%}" for d in wacc_deltas],
        columns=[f"{g:.1%}" for g in g_values],
    )
    st.dataframe(sens_df.style.format("{:.2f}").background_gradient(cmap="RdYlBu_r", axis=None),
                 use_container_width=True)

    if company["news"]:
        st.subheader("Actualités récentes")
        for title in company["news"]:
            st.write(f"- {title}")

    excel_buf = build_excel(company, wacc_result, fcf_df, dcf_result, sens_grid, wacc_deltas, g_values)
    st.download_button(
        "📥 Télécharger le modèle Excel",
        data=excel_buf,
        file_name=f"{ticker}_DCF_model.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

else:
    st.info("Renseigne un ticker dans la barre latérale et clique sur \"Lancer l'analyse\".")
